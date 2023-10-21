import datetime
from tkinter import filedialog

import openpyxl
from docxtpl import DocxTemplate

import db
import sys
from prettytable import PrettyTable
from operator import attrgetter
import create_log
import db


# управление чеками
def check_managment():
    def view_services():
        service_table = PrettyTable()
        service_table.field_names = ["id", "Название", "Стоимость"]
        for service in db.services:
            service_table.add_row([service.id, service.name, service.cost])
        print(service_table)

    def create_checks():
        def check_date_valid():
            try:
                date = datetime.datetime.strptime(check_date, "%d/%m/%Y")
                return True
            except Exception:
                print("Дата имеет неверный формат!")
                return False

        all_services = []
        check_user = create_log.logged_user
        if len(db.checks) != 0:
            check_id = int(db.checks[len(db.checks) - 1].id) + 1
        else:
            check_id = 1
        check_address = input('Введите адрес в формате: "[название улицы],[номер дома],кв.[номер квартиры]"')
        date_is_valid = False
        while not date_is_valid:
            check_date = input('Введите дату в формате: "дд/мм/гг"')
            date_is_valid = check_date_valid()

        def new_dialog():
            def check_service():
                for service in db.services:
                    if str(service.id) == id_service:
                        return True
                print("Услуга с таким id не существует!")
                return False

            def get_service():
                for service in db.services:
                    if str(service.id) == id_service:
                        return service

            print("Добавить новую услугу в чек?\n"
                  "1 - Да\n"
                  "2 - Нет\n")
            answer = input()
            if answer == '1':
                view_services()
                service_valid = False
                while not service_valid:
                    id_service = input("Введите id  услуги, которую хотите добавить:")
                    service_valid = check_service()
                service = get_service()
                all_services.append(service)
                new_dialog()
            if answer == '2':
                check_services = []
                [check_services.append(x) for x in all_services if x not in check_services]
                check_sum = 0
                for service in check_services:
                    check_sum += int(service.cost)
                new_check = db.Check(check_id, check_user, check_address, check_date, check_sum, check_services)
                db.checks.append(new_check)
                db.update_data()
                db.upload_data(db.database)

        new_dialog()

    def view_checks():
        check_table = PrettyTable()
        check_table.field_names = ["id", "Дата", "ФИО сотрудника", "Заказчик", "Услуги", "Сумма к оплате"]
        for check in db.checks:
            all_services = ''
            for service in check.services:
                all_services = f'{all_services}, {service.name}'
            all_services = all_services.replace(',', '', 1)
            check_table.add_row(
                [check.id, check.date, check.user.get_fullname(), check.address, all_services, check.sum])
        print(check_table)

    def dialog():
        print("\tУправление чеками\n"
              "[1] - Просмотреть чеки\n"
              "[2] - Создать чек\n"
              "[0] - Назад")
        result = input()
        if result == '1':
            view_checks()
            dialog()
        elif result == '2':
            create_checks()
            dialog()
        elif result == '0':
            return
        else:
            dialog()

    dialog()


# создание документов
def document_create():
    def view_checks():
        check_table = PrettyTable()
        check_table.field_names = ["id", "Дата", "ФИО сотрудника", "Заказчик", "Услуги", "Сумма к оплате"]
        for check in db.checks:
            all_services = ''
            for service in check.services:
                all_services = f'{all_services}, {service.name}'
            all_services = all_services.replace(',', '', 1)
            check_table.add_row(
                [check.id, check.date, check.user.get_fullname(), check.address, all_services, check.sum])
        print(check_table)

    def word_create(check):
        doc = DocxTemplate("example.docx")
        list = []

        for service in check.services:
            dic = {'cols': [service.name, service.id, service.cost]}
            list.append(dic)
        path = filedialog.askdirectory()
        file_name = path + '/' + check.user.middle_name + '_' + datetime.datetime.now().strftime('%d.%m.%y') + '.docx'
        context = {'num': check.id, 'user': check.user.get_fullname(), 'address': check.address,
                   'date': check.date.replace('/', '.'), 'sum': check.sum,
                   'col_labels': ['Наименование услуги', 'Артикул', 'Цена'], 'tbl_contents': list}
        doc.render(context)
        doc.save(file_name)

        print(f"Файл успешно сохранен! {file_name}")

    def excel_create(check):
        def dict_create(services):
            dict_services = []
            for service in services:
                dict_service = []
                dict_service.append(service.name)
                dict_service.append(str(service.id))
                dict_service.append(str(service.cost))
                dict_services.append(dict_service)
            return dict_services

        all_data = dict_create(check.services)

        path = filedialog.askdirectory()
        file_name = path + '/' + check.user.middle_name + '_' + datetime.datetime.now().strftime('%d.%m.%y') + '.xlsx'
        wb = openpyxl.load_workbook('example.xlsx')
        ws = wb['Лист1']
        ws['C1'].value = check.id
        ws['E1'].value = check.date.replace('/', '.')
        ws['B3'].value = check.user.get_fullname()
        ws['B4'].value = check.address
        for data in all_data:
            ws.append(data)

        ws.append([])
        ws.append(['Итог', str(check.sum)])
        wb.close()
        wb.save(file_name)
        print(f"Файл успешно сохранен! {file_name}")

    def view_checks():
        check_table = PrettyTable()
        check_table.field_names = ["id", "Дата", "ФИО сотрудника", "Заказчик", "Услуги", "Сумма к оплате"]
        for check in db.checks:
            all_services = ''
            for service in check.services:
                all_services = f'{all_services}, {service.name}'
            all_services = all_services.replace(',', '', 1)
            check_table.add_row(
                [check.id, check.date, check.user.get_fullname(), check.address, all_services, check.sum])
        print(check_table)

    def dialog():
        def find_check():
            for check in db.checks:
                if check.id == check_id:
                    return check

        def check_checks():
            for check in db.checks:
                if check.id == check_id:
                    return True
            print("Чек с таким id не существует!")
            return False

        is_check_exist = False
        while not is_check_exist:
            check_id = input("Введите id чека: ")
            is_check_exist = check_checks()
        finded_check = find_check()
        return finded_check

    view_checks()
    check = dialog()

    def dialog_2():
        print("Выберите формат файла:\n"
              "[1] - .doc(word)\n"
              "[2] - .xlsx(excel)\n"
              "[0] - Назад")
        result = input()
        if result == '1':
            word_create(check)
        elif result == '2':
            excel_create(check)
        elif result == '0':
            return
        else:
            dialog_2()

    dialog_2()


# блок меню пользователя
def user_menu():
    def dialog():
        print("\tМеню пользователя\n"
              "[1] - Управление чеками\n"
              "[2] - Создать документ\n"
              "[0] - Выйти")
        result = input()
        if result == '1':
            check_managment()
            dialog()
        elif result == '2':
            document_create()
            dialog()
        elif result == '0':
            return
        else:
            dialog()

    dialog()
